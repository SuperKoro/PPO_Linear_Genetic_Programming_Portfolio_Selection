import openpyxl
import json
import os

# Tạo folder data nếu chưa tồn tại
os.makedirs('data', exist_ok=True)

# Đọc file Excel
excel_file = 'Excel Data_23_IELSIU20327_Trần Đức Khiêm_GVHD_Assoc.Prof. Nguyen Van Hop.xlsx'
wb = openpyxl.load_workbook(excel_file, read_only=False, data_only=True)

print(f"Found {len(wb.sheetnames)} sheets: {wb.sheetnames}\n")

# Chỉ xử lý các sheet Set (bỏ qua Validation)
target_sheets = ['Set20', 'Set25', 'Set30', 'Set35', 'Set40', 'Set45', 'Set50']

for sheet_name in target_sheets:
    if sheet_name not in wb.sheetnames:
        print(f"⚠ Sheet {sheet_name} not found, skipping...")
        continue
        
    print(f"Processing sheet: {sheet_name}")
    ws = wb[sheet_name]
    
    # In ra vài dòng đầu để debug
    print("  First few rows:")
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True)):
        print(f"    Row {idx+1}: {row}")
    
    # Đọc dữ liệu từ sheet
    jobs_data = {}
    machine_pool = set()
    
    # Giả sử format: Job | Operation | Machine1 | Machine2 | ... | ProcessingTime
    # Cần xác định cột nào là processing time (thường là cột cuối)
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or row[0] is None:  # Bỏ qua dòng trống
            continue
        
        try:
            job_id = int(row[0])  # Job ID
            op_id = int(row[1])   # Operation ID
        except (ValueError, TypeError):
            continue
        
        # Tìm processing time (cột cuối cùng có giá trị)
        processing_time = None
        candidate_machines = []
        
        # Duyệt từ cột 2 đến cuối
        for col_idx in range(2, len(row)):
            val = row[col_idx]
            if val is not None and val != '':
                try:
                    num_val = int(val)
                    # Nếu là cột cuối hoặc cột tiếp theo là None, đây là processing time
                    if col_idx == len(row) - 1 or (col_idx < len(row) - 1 and row[col_idx + 1] is None):
                        processing_time = num_val
                    else:
                        candidate_machines.append(num_val)
                        machine_pool.add(num_val)
                except (ValueError, TypeError):
                    pass
        
        # Tạo operation dictionary
        operation = {
            'op_id': op_id,
            'candidate_machines': candidate_machines,
            'processing_time': processing_time
        }
        
        # Thêm vào jobs_data
        if job_id not in jobs_data:
            jobs_data[job_id] = []
        jobs_data[job_id].append(operation)
    
    # Tạo dataset hoàn chỉnh
    dataset = {
        'name': sheet_name,
        'machine_pool': sorted(list(machine_pool)),
        'jobs': {str(k): v for k, v in sorted(jobs_data.items())},
        'due_dates': {str(i): 1200 for i in range(1, len(jobs_data) + 1)}
    }
    
    # Lưu vào file JSON
    output_file = f'data/{sheet_name}.json'
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(dataset, f, indent=2, ensure_ascii=False)
    
    print(f"✓ Saved {output_file}")
    print(f"  - Number of jobs: {len(jobs_data)}")
    print(f"  - Machine pool: {sorted(list(machine_pool))}")
    print()

print("✅ All datasets converted successfully!")
